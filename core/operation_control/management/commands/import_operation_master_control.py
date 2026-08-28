from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.operation_control.models import OperationMasterControl, OperationControlChangeLog
from core.operations_panel.models import Operation

# Columnas de la hoja 2026 (indices 0-based al usar values_only=True)
COL_CODIGO_PLANTILLA = 0          # A - CODIGO PLANTILLA
COL_CONTRARECIBO = 1              # B - CONTRARECIBO (sin campo equivalente en el modelo actual)
COL_FECHA_VIAJE = 2               # C - FECHA VIAJE (no se importa al control maestro)
COL_CODIGO_VIAJE = 3              # D - CODIGO VIAJE -> Operation.folio
COL_FOLIO_FACTURA_CLIENTE = 4     # E - FOLIO FACTURA
COL_FECHA_FACTURA_CLIENTE = 5     # F - FECHA FACTURA
COL_PRECIO = 12                   # M - PRECIO
COL_FECHA_COBRO = 15              # P - FECHA COBRO
COL_COSTO = 16                    # Q - COSTO
COL_FECHA_FACTURA_PROVEEDOR = 21  # V - FECHA FACTURA
COL_FOLIO_FACTURA_PROVEEDOR = 22  # W - FOLIO FACTURA PROVEEDOR
COL_FECHA_PROGRAMADA_PAGO = 23    # X - FECHA PROGRAMADA PAGO

MAX_COL = 24


def clean_text(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def clean_decimal(value):
    if value is None or value == "":
        return None

    if isinstance(value, str):
        value = value.strip()
        if not value or value.upper() in {"NA", "N/A", "NONE", "NULL", "-"}:
            return None
        value = value.replace("$", "").replace(",", "")

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError, ValueError):
        return None


def clean_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        value = value.strip()
        if not value or value.upper() in {"NA", "N/A", "NONE", "NULL", "-"}:
            return None

        formats = (
            "%d/%m/%Y",
            "%d/%m/%y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d-%m-%y",
        )

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass

    return None


def display_value(value):
    if value is None:
        return ""
    return str(value)


class Command(BaseCommand):
    help = (
        "Importa CONTROL GENERAL MAESTRO 2026.xlsx y crea/actualiza "
        "OperationMasterControl buscando Operation por folio = CODIGO VIAJE."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Ruta del archivo Excel a importar.",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="2026",
            help="Nombre de la hoja. Default: 2026",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Valida y muestra resultados sin guardar cambios.",
        )
        parser.add_argument(
            "--skip-existing",
            action="store_true",
            help="No modifica controles maestros que ya existan.",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]
        skip_existing = options["skip_existing"]

        if not excel_path.exists():
            raise CommandError(f"No existe el archivo: {excel_path}")

        # Evita depender de la ruta Python exacta de los modelos.


        self.stdout.write(f"Abriendo: {excel_path}")
        self.stdout.write(f"Hoja: {sheet_name}")

        # data_only=True es importante: COSTO y FECHA PROGRAMADA PAGO
        # contienen formulas en el Excel y queremos su valor calculado.
        workbook = load_workbook(
            filename=excel_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"No existe la hoja '{sheet_name}'. "
                f"Hojas disponibles: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]
        self.validate_headers(worksheet)

        rows = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=MAX_COL, values_only=True),
            start=2,
        ):
            folio = clean_text(row[COL_CODIGO_VIAJE])
            if not folio:
                continue

            rows.append((row_number, folio, row))

        if not rows:
            self.stdout.write(self.style.WARNING("No se encontraron filas para importar."))
            return

        # El OneToOne contra Operation implica que dos filas con el mismo codigo
        # de viaje no pueden importarse con seguridad. Se omiten para revision.
        excel_counts = Counter(folio for _, folio, _ in rows)
        duplicated_excel_folios = {
            folio for folio, count in excel_counts.items() if count > 1
        }

        safe_rows = [
            item for item in rows if item[1] not in duplicated_excel_folios
        ]
        safe_folios = [folio for _, folio, _ in safe_rows]

        # Carga las operaciones de una sola vez para evitar ~3000 consultas.
        operations = list(Operation.objects.filter(folio__in=safe_folios))

        db_counts = Counter(clean_text(operation.folio) for operation in operations)
        duplicated_db_folios = {
            folio for folio, count in db_counts.items() if count > 1
        }

        operations_by_folio = {
            clean_text(operation.folio): operation
            for operation in operations
            if clean_text(operation.folio) not in duplicated_db_folios
        }

        operation_ids = [operation.pk for operation in operations_by_folio.values()]
        existing_controls = {
            control.operation_id: control
            for control in OperationMasterControl.objects.filter(
                operation_id__in=operation_ids
            )
        }

        created = 0
        updated = 0
        unchanged = 0
        skipped_existing = 0
        missing_operations = []
        parse_warnings = []

        def execute_import():
            nonlocal created, updated, unchanged, skipped_existing

            for row_number, folio, row in safe_rows:
                operation = operations_by_folio.get(folio)

                if operation is None:
                    missing_operations.append((row_number, folio))
                    continue

                values, warnings = self.build_values(row_number, row)
                parse_warnings.extend(warnings)

                control = existing_controls.get(operation.pk)

                if control is None:
                    if not dry_run:
                        control = OperationMasterControl.objects.create(
                            operation=operation,
                            **values,
                        )
                        existing_controls[operation.pk] = control
                    created += 1
                    continue

                if skip_existing:
                    skipped_existing += 1
                    continue

                changed_fields = []
                change_logs = []

                for field_name, new_value in values.items():
                    old_value = getattr(control, field_name)

                    if old_value == new_value:
                        continue

                    changed_fields.append(field_name)

                    if not dry_run:
                        setattr(control, field_name, new_value)
                        change_logs.append(
                            OperationControlChangeLog(
                                control=control,
                                field_name=field_name,
                                previous_value=display_value(old_value),
                                new_value=display_value(new_value),
                                changed_by=None,
                            )
                        )

                if not changed_fields:
                    unchanged += 1
                    continue

                if not dry_run:
                    # updated_at es auto_now y debe incluirse al usar update_fields.
                    control.save(update_fields=[*changed_fields, "updated_at"])
                    if change_logs:
                        OperationControlChangeLog.objects.bulk_create(change_logs)

                updated += 1

        if dry_run:
            execute_import()
        else:
            # Todo el proceso se confirma o revierte como una sola transaccion.
            with transaction.atomic():
                execute_import()

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Importacion finalizada."))
        self.stdout.write(f"Filas con CODIGO VIAJE: {len(rows)}")
        self.stdout.write(f"Folios unicos en Excel: {len(excel_counts)}")
        self.stdout.write(f"Creados: {created}")
        self.stdout.write(f"Actualizados: {updated}")
        self.stdout.write(f"Sin cambios: {unchanged}")
        self.stdout.write(f"Existentes omitidos: {skipped_existing}")
        self.stdout.write(f"Operacion no encontrada: {len(missing_operations)}")
        self.stdout.write(
            f"Codigos duplicados en Excel: {len(duplicated_excel_folios)}"
        )
        self.stdout.write(
            f"Folios duplicados en BD: {len(duplicated_db_folios)}"
        )
        self.stdout.write(f"Advertencias de conversion: {len(parse_warnings)}")

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN: no se guardo ningun cambio."))

        if duplicated_excel_folios:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Duplicados en el Excel (omitidos):"))
            for folio in sorted(duplicated_excel_folios):
                excel_rows = [
                    str(row_number)
                    for row_number, row_folio, _ in rows
                    if row_folio == folio
                ]
                self.stdout.write(f"  - {folio}: filas {', '.join(excel_rows)}")

        if duplicated_db_folios:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Folios duplicados en BD (omitidos):"))
            for folio in sorted(duplicated_db_folios):
                self.stdout.write(f"  - {folio}")

        if missing_operations:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Primeras operaciones no encontradas:"))
            for row_number, folio in missing_operations[:50]:
                self.stdout.write(f"  - Fila {row_number}: {folio}")
            if len(missing_operations) > 50:
                self.stdout.write(
                    f"  ... y {len(missing_operations) - 50} adicionales."
                )

        if parse_warnings:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Primeras advertencias de conversion:"))
            for warning in parse_warnings[:50]:
                self.stdout.write(f"  - {warning}")
            if len(parse_warnings) > 50:
                self.stdout.write(
                    f"  ... y {len(parse_warnings) - 50} adicionales."
                )

    def validate_headers(self, worksheet):
        expected = {
            1: "CODIGO PLANTILLA",
            2: "CONTRARECIBO",
            4: "CODIGO VIAJE",
            5: "FOLIO FACTURA",
            6: "FECHA FACTURA",
            13: "PRECIO",
            16: "FECHA COBRO",
            17: "COSTO",
            22: "FECHA FACTURA",
            23: "FOLIO FACTURA PROVEEDOR",
            24: "FECHA PROGRAMADA PAGO",
        }

        errors = []
        for column, expected_name in expected.items():
            actual_name = clean_text(worksheet.cell(row=1, column=column).value).upper()
            if actual_name != expected_name:
                errors.append(
                    f"Columna {column}: esperaba '{expected_name}' y encontro "
                    f"'{actual_name or '<vacio>'}'"
                )

        if errors:
            raise CommandError(
                "La estructura del Excel no coincide con la esperada:\n- "
                + "\n- ".join(errors)
            )

    def build_values(self, row_number, row):
        warnings = []
        values = {}

        codigo_plantilla = clean_text(row[COL_CODIGO_PLANTILLA])
        if codigo_plantilla:
            # En tu modelo counter_receipt tiene verbose_name='Codigo de plantilla'.
            values["counter_receipt"] = codigo_plantilla

        customer_invoice_code = clean_text(row[COL_FOLIO_FACTURA_CLIENTE])
        if customer_invoice_code and customer_invoice_code.upper() not in {"NA", "N/A"}:
            values["customer_invoice_code"] = customer_invoice_code

        customer_invoice_date = clean_date(row[COL_FECHA_FACTURA_CLIENTE])
        if customer_invoice_date is not None:
            values["customer_invoice_date"] = customer_invoice_date
        elif row[COL_FECHA_FACTURA_CLIENTE] not in (None, ""):
            raw = clean_text(row[COL_FECHA_FACTURA_CLIENTE])
            if raw.upper() not in {"NA", "N/A", "NONE", "NULL", "-"}:
                warnings.append(
                    f"Fila {row_number}: FECHA FACTURA CLIENTE no reconocida: {raw}"
                )

        expected_collection_date = clean_date(row[COL_FECHA_COBRO])
        if expected_collection_date is not None:
            values["expected_collection_date"] = expected_collection_date
        elif row[COL_FECHA_COBRO] not in (None, ""):
            warnings.append(
                f"Fila {row_number}: FECHA COBRO no reconocida: "
                f"{clean_text(row[COL_FECHA_COBRO])}"
            )

        sale_amount = clean_decimal(row[COL_PRECIO])
        if sale_amount is not None:
            values["sale_amount_override"] = sale_amount
        elif row[COL_PRECIO] not in (None, ""):
            warnings.append(
                f"Fila {row_number}: PRECIO no reconocido: {clean_text(row[COL_PRECIO])}"
            )

        cost_amount = clean_decimal(row[COL_COSTO])
        if cost_amount is not None:
            values["cost_amount_override"] = cost_amount
        elif row[COL_COSTO] not in (None, ""):
            warnings.append(
                f"Fila {row_number}: COSTO no reconocido: {clean_text(row[COL_COSTO])}"
            )

        supplier_invoice_date = clean_date(row[COL_FECHA_FACTURA_PROVEEDOR])
        if supplier_invoice_date is not None:
            values["supplier_invoice_date"] = supplier_invoice_date
        elif row[COL_FECHA_FACTURA_PROVEEDOR] not in (None, ""):
            raw = clean_text(row[COL_FECHA_FACTURA_PROVEEDOR])
            if raw.upper() not in {"NA", "N/A", "NONE", "NULL", "-"}:
                warnings.append(
                    f"Fila {row_number}: FECHA FACTURA PROVEEDOR no reconocida: {raw}"
                )

        supplier_invoice_number = clean_text(row[COL_FOLIO_FACTURA_PROVEEDOR])
        if supplier_invoice_number and supplier_invoice_number.upper() not in {"NA", "N/A"}:
            values["supplier_invoice_number"] = supplier_invoice_number

        scheduled_payment_date = clean_date(row[COL_FECHA_PROGRAMADA_PAGO])
        if scheduled_payment_date is not None:
            values["scheduled_supplier_payment_date"] = scheduled_payment_date
        elif row[COL_FECHA_PROGRAMADA_PAGO] not in (None, ""):
            warnings.append(
                f"Fila {row_number}: FECHA PROGRAMADA PAGO no reconocida: "
                f"{clean_text(row[COL_FECHA_PROGRAMADA_PAGO])}"
            )

        return values, warnings
