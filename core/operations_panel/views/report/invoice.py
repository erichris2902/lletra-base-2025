from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import timedelta, datetime

from apps.facturapi.models import FacturapiInvoice
from core.operations_panel.choices import OperationStatus
from core.operations_panel.models import Operation




def report_xml_invoices(request):
    fecha_inicio_str = request.POST.get("fecha_inicial")
    fecha_fin_str = request.POST.get("fecha_final")

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Faltan parámetros: fecha_inicio y fecha_fin", status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido. Usa YYYY-MM-DD.", status=400)

    facturas = (
        FacturapiInvoice.objects.filter(stamp_date__range=[fecha_inicio, fecha_fin])
        .select_related()
        .order_by("stamp_date")
    )

    if not facturas.exists():
        return HttpResponse("No se encontraron facturas en el rango indicado.", status=404)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = [
        "FECHA",
        "FECHA DE CARGA",
        "CONTROL VEHICULAR",
        "FOLIO",
        "FOLIO FISCAL",
        "STATUS",
        "RECEPTOR",
        "TOTAL (CON IMPUESTOS)",
        "TIPO DE CFDI",
        "SERVICIO",
        "KMs",
        "ORIGEN (COLONIA)",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Recorremos cada factura
    for inv in facturas:
        # Buscar operación asociada (por shipment_invoice o M2M)
        operation = (
                Operation.objects.filter(shipment_invoice=inv).first()
                or Operation.objects.filter(invoices=inv).first()
        )

        fecha_operacion = operation.operation_date if operation else None
        fecha_carga = operation.cargo_appointment if operation and operation.cargo_appointment else None
        control_vehicular = getattr(operation.vehicle, "plate", "") if operation and operation.vehicle else ""
        servicio = operation.get_shipment_type_display() if operation else ""
        kms = getattr(operation.route, "direct_distance", "") if operation and operation.route else ""
        origen = getattr(operation.route.initial_location.address, "colony", "") if operation and operation.route and operation.route.initial_location and operation.route.initial_location.address else ""

        receptor = ""
        if inv.customer:
            receptor = inv.customer or ""

        fila = [
            operation.operation_date.strftime("%d/%m/%Y") if operation else "",
            fecha_carga.strftime("%d/%m/%Y") if fecha_carga else "",
            operation.folio if operation else "",
            str(inv.series) + str(inv.folio_number),
            inv.uuid,
            inv.status or "",
            str(receptor),
            float(inv.total) if inv.total else "",
            inv.type or "",
            servicio,
            kms,
            origen,
        ]
        ws.append(fila)

    # Ajustar columnas automáticamente
    for c in range(1, len(headers) + 1):
        max_len = len(headers[c - 1])
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            max_len = max(max_len, len(str(val)) if val else 0)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 45)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"

    # Guardar archivo en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"facturas_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
