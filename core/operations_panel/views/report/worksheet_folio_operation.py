from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import timedelta, datetime
from core.operations_panel.choices import OperationStatus
from core.operations_panel.models import Operation


def _agregar_encabezados(ws, blueFill):
    headers = [
        "FECHA", "VIAJE", "SERVICIO", "CLIENTE", "ORIGEN", "DESTINO", "REPARTOS",
        "UNIDAD", "TARIFA", "OPERADOR", "PROVEEDOR", "CONCEPTO", "PACKING", "FACTURADO"
    ]
    ws.append(headers)
    for col in ws[1]:
        col.fill = blueFill
        col.font = Font(color="FFFFFF")


def _agregar_fila(op, ws, redFill, greenFill):
    print(op.supplier)
    date = op.raw_payload.get('fecha', '') if op.raw_payload else ''
    client = op.raw_payload.get('cliente', '') if op.raw_payload else ''
    origin = op.raw_payload.get('origen', '') if op.raw_payload else ''
    destiny = op.raw_payload.get('destino', '') if op.raw_payload else ''
    stops = op.raw_payload.get('repartos', '') if op.raw_payload else ''
    unit = op.raw_payload.get('unidad', '') if op.raw_payload else ''
    driver = op.raw_payload.get('operador', '') if op.raw_payload else ''
    supplier = op.raw_payload.get('proveedor', '') if op.raw_payload else ''

    invoiced = "Sí" if op.shipment_invoice else "No"
    packing = "Sí" if op.transported_products.count() > 0 else "No"
    row = [
        date,
        str(op.folio or ""),
        "TRASLADO",
        client,
        origin,
        destiny,
        str(stops),
        unit,
        str(op.shipment_invoice.total if op.shipment_invoice else "0.00"),
        driver,
        supplier,
        f"{op.folio} Ruta: {op.route.destination_location}, {op.operation_date}" if op.route else "",
        packing,
        invoiced,
    ]
    ws.append(row)
    r = ws.max_row

    if op.status == OperationStatus.CANCELLED:
        for c in range(1, len(row) + 1):
            ws.cell(row=r, column=c).fill = redFill
        return

    # Facturado
    ws.cell(row=r, column=14).fill = (
        greenFill if op.shipment_invoice else redFill
    )

    # Packing
    ws.cell(row=r, column=13).fill = (
        greenFill if op.transported_products.count() > 0 else redFill
    )


def _ajustar_columnas(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    dims.get(cell.column_letter, 0), len(str(cell.value)) + 5
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = min(value, 50)

def report_xml_worksheet_folios_by_date(request):
    fecha_inicio_str = request.POST.get("fecha_inicial")
    fecha_fin_str = request.POST.get("fecha_final")

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Faltan parámetros: fecha_inicio y fecha_fin", status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido. Usa YYYY-MM-DD.", status=400)

    operations = Operation.objects.filter(
        operation_date__range=[fecha_inicio, fecha_fin]
    ).order_by("operation_date")

    if not operations.exists():
        return HttpResponse("No hay operaciones en el rango seleccionado.", status=404)

    # Crear libro de Excel
    wb = Workbook()

    # Estilos
    blueFill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    greenFill = PatternFill(start_color='00EF23', end_color='00EF23', fill_type='solid')
    redFill = PatternFill(start_color='ED000B', end_color='ED000B', fill_type='solid')

    # Agrupar operaciones por fecha
    grouped = {}
    for op in operations:
        grouped.setdefault(op.operation_date, []).append(op)

    # Eliminar hoja por defecto si hay más de una fecha
    if len(grouped) > 1 and wb.active.title == "Sheet":
        wb.remove(wb.active)

    for fecha, ops in grouped.items():
        ws = wb.create_sheet(title=str(fecha))
        _agregar_encabezados(ws, blueFill)

        for op in ops:
            _agregar_fila(op, ws, redFill, greenFill)

        # Ajustar anchos automáticamente
        _ajustar_columnas(ws)

    # Si solo hay una hoja, renombrarla
    if len(grouped) == 1:
        ws = wb.active
        ws.title = str(list(grouped.keys())[0])

    # Guardar Excel en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"hoja_trabajo_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def generar_reporte_hoja_trabajo(fecha_inicio, fecha_fin):
    operations = Operation.objects.filter(
        operation_date__range=[fecha_inicio, fecha_fin]
    ).order_by("operation_date")

    headers = [
        "Fecha",
        "Viaje",
        "Servicio",
        "Cliente",
        "Origen",
        "Destino",
        "Repartos",
        "Unidad",
        "Tarifa (Con IVA)",
        "Operador",
        "Proveedor",
        "Concepto",
        "Packing",
        "Facturado",
    ]

    rows = []
    for op in operations:
        supplier = getattr(op.supplier, "name", "")
        invoiced = "Sí" if op.shipment_invoice else "No"
        packing = "Sí" if op.transported_products.count() > 0 else "No"

        rows.append([
            op.operation_date.strftime("%d/%m/%Y"),
            str(op.folio or ""),
            "TRASLADO",
            str(op.client or ""),
            str(op.route.initial_location if op.route else ""),
            str(op.route.destination_location if op.route else ""),
            str(op.route.route_stops.count() if op.route else ""),
            str(op.vehicle or ""),
            str(op.shipment_invoice.total if op.shipment_invoice else "-"),
            str(op.driver or ""),
            supplier,
            f"{op.folio} Ruta: {op.route.destination_location}, {op.operation_date}" if op.route else "",
            packing,
            invoiced,
        ])

    return headers, rows

def report_xml_worksheet_folios_by_folio(request):
    folio_serie = request.POST.get("folio_serie")
    folio_number = request.POST.get("folio_number")

    operations = Operation.objects.filter(
        vehicle__operation__folio__startswith=str(folio_serie),
    ).order_by("-folio")

    if not operations.exists():
        return HttpResponse("No hay operaciones en el rango seleccionado.", status=404)

    # Crear libro de Excel
    wb = Workbook()

    # Estilos
    blueFill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    greenFill = PatternFill(start_color='00EF23', end_color='00EF23', fill_type='solid')
    redFill = PatternFill(start_color='ED000B', end_color='ED000B', fill_type='solid')

    # Agrupar operaciones por fecha
    grouped = {}
    for op in operations:
        print(op.folio)
        grouped.setdefault(op.operation_date, []).append(op)

    # Eliminar hoja por defecto si hay más de una fecha
    if len(grouped) > 1 and wb.active.title == "Sheet":
        wb.remove(wb.active)

    for fecha, ops in grouped.items():
        ws = wb.create_sheet(title=str(fecha))
        _agregar_encabezados(ws, blueFill)

        for op in ops:
            _agregar_fila(op, ws, redFill, greenFill)

        # Ajustar anchos automáticamente
        _ajustar_columnas(ws)

    # Si solo hay una hoja, renombrarla
    if len(grouped) == 1:
        ws = wb.active
        ws.title = str(list(grouped.keys())[0])

    # Guardar Excel en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response





