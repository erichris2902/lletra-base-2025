from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

def create_excel_response(headers, rows, filename_prefix="reporte", status_styles=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Estilos básicos
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Escribir encabezados
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Escribir filas
    for row in rows:
        ws.append(row)

    # Congelar encabezados
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Auto ancho de columnas
    for c in range(1, len(headers) + 1):
        max_len = len(str(headers[c - 1])) if headers[c - 1] else 0
        for r in range(2, len(rows) + 2):
            val = ws.cell(row=r, column=c).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 50)

    # Estilos condicionales opcionales
    if status_styles:
        status_col = headers.index("Estatus") + 1 if "Estatus" in headers else None
        if status_col:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=status_col)
                key = str(cell.value or "").strip().lower()
                fill_rgb, font_rgb = status_styles.get(
                    key, ("FFEEEEEE", "FF000000")
                )
                cell.fill = PatternFill(fill_type="solid", start_color=fill_rgb, end_color=fill_rgb)
                cell.font = Font(color=font_rgb, bold=True)

    # Generar respuesta HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    from django.http import HttpResponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
