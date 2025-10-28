
from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.rh_panel.models import Attendance


def report_attendance(request):
    fecha_inicio_str = request.POST.get("fecha_inicial")
    fecha_fin_str = request.POST.get("fecha_final")

    if not fecha_inicio_str or not fecha_fin_str:
        return HttpResponse("Faltan parámetros: fecha_inicial y fecha_final", status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido. Usa YYYY-MM-DD.", status=400)

    asistencias = (
        Attendance.objects
        .select_related("employee")
        .filter(date__range=[fecha_inicio, fecha_fin])
        .order_by("date", "employee__name")
    )

    if not asistencias.exists():
        return HttpResponse("No se encontraron asistencias en el rango indicado.", status=404)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    headers = [
        "FECHA",
        "EMPLEADO",
        "ENTRADA",
        "SALIDA",
        "CONFIANZA",
        "EMOCIÓN ENTRADA",
        "EMOCIÓN SALIDA",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Recorremos cada registro
    for a in asistencias:
        fila = [
            a.date.strftime("%d/%m/%Y") if a.date else "",
            str(a.employee) if a.employee else "",
            a.check_in.strftime("%H:%M") if a.check_in else "",
            a.check_out.strftime("%H:%M") if a.check_out else "",
            round(a.confidence, 2),
            a.emotion_check_in or "",
            a.emotion_check_out or "",
        ]
        ws.append(fila)

    # Ajustar columnas automáticamente
    for c in range(1, len(headers) + 1):
        max_len = len(headers[c - 1])
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            max_len = max(max_len, len(str(val)) if val else 0)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 40)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"

    # Guardar archivo en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"asistencias_{fecha_inicio}_a_{fecha_fin}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
