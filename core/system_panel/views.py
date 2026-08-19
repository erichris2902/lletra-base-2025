from datetime import datetime
from decimal import Decimal, InvalidOperation

import dateutil
from dateutil.utils import today
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import HttpResponseBadRequest, HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.borders import BORDER_THIN

from apps.facturapi.models import FacturapiProduct, FacturapiTax, FacturapiInvoice
from core.operations_panel.choices import AsturianoPacking
from core.operations_panel.models import Operation, TransportedProduct, Client
from core.operations_panel.models.distribution_packing import DistributionPacking
from core.operations_panel.views.report.attendance import report_attendance
from core.operations_panel.views.report.invoice import report_xml_invoices
from core.operations_panel.views.report.worksheet_folio_operation import report_xml_worksheet_folios_by_date, \
    report_xml_worksheet_folios_by_folio, report_xml_worksheet_folios_by_date2
from core.system.models import Category, Section
from core.system.views import AdminTemplateView, AdminListView
from core.system_panel.forms import CategoryForm, SectionForm, AssistantForm, ActionEngineForm, ReportEngineForm, \
    ReportEngineByFolioForm, ExpedienteZipForm, PdfReduceZipForm
from apps.openai_assistant.models import Assistant
from core.operation_control.models import OperationMasterControl
from core.admin_panel.models.purchase_order import PurchaseOrderOperation, PurchaseOrderAccessory, PurchaseOrderStatus
from django.db.models import Prefetch


class DashboardView(LoginRequiredMixin, AdminTemplateView):
    """
    Dashboard view for SYSTEM users.
    """
    template_name = 'system_panel/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dashboard de Sistema'
        return context

class CategoryListView(AdminListView):
    model = Category
    form = CategoryForm
    template_name = 'base/elements/views/datatable_list.html'
    datatable_headers = ["Nombre", "Prioridad", "URL", "Sistema"]
    datatable_keys = ["name", "priority", "url", "system"]
    datatable_actions = True
    title = model._meta.verbose_name_plural.title()
    form_path = 'base/elements/forms/form.html'
    section = 'Categorias de las barras de navegacion'
    category = 'Sidebar'

class SectionListView(AdminListView):
    model = Section
    form = SectionForm
    template_name = 'base/elements/views/datatable_list.html'
    datatable_headers = ["Nombre", "Prioridad", "URL", "Categoria", "Activo"]
    datatable_keys = ["name", "priority", "url", "category", "is_active"]
    datatable_actions = True
    title = model._meta.verbose_name_plural.title()
    form_path = 'base/elements/forms/form.html'
    section = 'Secciones de las barras de navegacion'
    category = 'Sidebar'

class AssistantListView(AdminListView):
    model = Assistant
    form = AssistantForm
    template_name = 'base/elements/views/datatable_list.html'
    datatable_headers = ["Nombre", "Command", "Modelo", "Activo"]
    datatable_keys = ["name", "telegram_command", "model", "is_active"]
    datatable_actions = True
    title = model._meta.verbose_name_plural.title()
    form_path = 'base/elements/forms/form.html'
    section = 'Asistentes de IA'
    category = 'OpenAI'


class ActionEngineView(AdminTemplateView):
    template_name = 'system_panel/actions_form.html'

    form_action = "ExecuteActionEngine"
    form_type = "vertical"
    title = "Motor de acciones"
    section = "Motor de acciones"
    category = "Facturacion MX"

    def post(self, request, *args, **kwargs):
        print(request.POST)
        print(request.FILES)
        action = request.POST.get("action")
        plantilla_excel = request.FILES.get("file")

        if not plantilla_excel:
            return HttpResponseBadRequest("Faltan incluir plantilla")
        print(action)
        if action == "CMR":
            try:
                workbook = load_workbook(
                    filename=plantilla_excel,
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.active
            except Exception as exc:
                return HttpResponseBadRequest(
                    f"No fue posible leer el archivo Excel: {exc}"
                )

            encabezados_requeridos = {
                "UUID A CANCELAR",
            }

            filas = worksheet.iter_rows(values_only=True)

            try:
                primera_fila = next(filas)
            except StopIteration:
                workbook.close()
                return HttpResponseBadRequest("El archivo está vacío.")

            encabezados = {
                str(valor).strip().upper(): indice
                for indice, valor in enumerate(primera_fila)
                if valor is not None
            }

            encabezados_faltantes = encabezados_requeridos - set(encabezados.keys())

            if encabezados_faltantes:
                workbook.close()
                return HttpResponseBadRequest(
                    "Faltan los siguientes encabezados: "
                    + ", ".join(sorted(encabezados_faltantes))
                )

            facturas_canceladas = 0
            filas_omitidas = []
            errores = []

            try:
                indice_uuid = encabezados["UUID A CANCELAR"]
                indice_uuid_relacion = encabezados["UUID A RELACIONAR"]

                with transaction.atomic():
                    for numero_fila, fila in enumerate(filas, start=2):

                        if indice_uuid >= len(fila):
                            filas_omitidas.append(
                                f"Fila {numero_fila}: no contiene la columna UUID A CANCELAR."
                            )
                            continue

                        uuid_a_cancelar = fila[indice_uuid]
                        uuid_a_relacionar = fila[indice_uuid_relacion]

                        if uuid_a_cancelar in (None, ""):
                            filas_omitidas.append(
                                f"Fila {numero_fila}: UUID vacío."
                            )
                            continue

                        uuid_a_cancelar = str(uuid_a_cancelar).strip()
                        uuid_a_relacionar = str(uuid_a_relacionar).strip()

                        try:
                            factura = FacturapiInvoice.objects.get(
                                uuid__iexact=uuid_a_cancelar
                            )
                        except FacturapiInvoice.DoesNotExist:
                            errores.append(
                                f"Fila {numero_fila}: no existe una factura "
                                f"con UUID {uuid_a_cancelar}."
                            )
                            continue

                        try:
                            factura.cancel(motivo="01", sustitucion=uuid_a_relacionar)
                            facturas_canceladas += 1
                        except Exception as exc:
                            errores.append(
                                f"Fila {numero_fila}: error al cancelar "
                                f"{uuid_a_cancelar}: {exc}"
                            )

            finally:
                workbook.close()

            mensaje = (
                f"Proceso terminado. "
                f"Facturas canceladas: {facturas_canceladas}. "
                f"Filas omitidas: {len(filas_omitidas)}. "
                f"Errores: {len(errores)}."
            )

            if errores:
                mensaje += "\n\n" + "\n".join(errores)

            return HttpResponse(mensaje, content_type="text/plain")

        if action == "CM":
            try:
                workbook = load_workbook(
                    filename=plantilla_excel,
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.active
            except Exception as exc:
                return HttpResponseBadRequest(
                    f"No fue posible leer el archivo Excel: {exc}"
                )

            encabezados_requeridos = {
                "UUID A CANCELAR",
            }

            filas = worksheet.iter_rows(values_only=True)

            try:
                primera_fila = next(filas)
            except StopIteration:
                workbook.close()
                return HttpResponseBadRequest("El archivo está vacío.")

            encabezados = {
                str(valor).strip().upper(): indice
                for indice, valor in enumerate(primera_fila)
                if valor is not None
            }

            encabezados_faltantes = encabezados_requeridos - set(encabezados.keys())

            if encabezados_faltantes:
                workbook.close()
                return HttpResponseBadRequest(
                    "Faltan los siguientes encabezados: "
                    + ", ".join(sorted(encabezados_faltantes))
                )

            facturas_canceladas = 0
            filas_omitidas = []
            errores = []

            try:
                indice_uuid = encabezados["UUID A CANCELAR"]

                with transaction.atomic():
                    for numero_fila, fila in enumerate(filas, start=2):

                        if indice_uuid >= len(fila):
                            filas_omitidas.append(
                                f"Fila {numero_fila}: no contiene la columna UUID A CANCELAR."
                            )
                            continue

                        uuid_a_cancelar = fila[indice_uuid]

                        if uuid_a_cancelar in (None, ""):
                            filas_omitidas.append(
                                f"Fila {numero_fila}: UUID vacío."
                            )
                            continue

                        uuid_a_cancelar = str(uuid_a_cancelar).strip()

                        try:
                            factura = FacturapiInvoice.objects.get(
                                uuid__iexact=uuid_a_cancelar
                            )
                        except FacturapiInvoice.DoesNotExist:
                            errores.append(
                                f"Fila {numero_fila}: no existe una factura "
                                f"con UUID {uuid_a_cancelar}."
                            )
                            continue

                        try:
                            factura.cancel("02")
                            facturas_canceladas += 1
                        except Exception as exc:
                            errores.append(
                                f"Fila {numero_fila}: error al cancelar "
                                f"{uuid_a_cancelar}: {exc}"
                            )

            finally:
                workbook.close()

            mensaje = (
                f"Proceso terminado. "
                f"Facturas canceladas: {facturas_canceladas}. "
                f"Filas omitidas: {len(filas_omitidas)}. "
                f"Errores: {len(errores)}."
            )

            if errores:
                mensaje += "\n\n" + "\n".join(errores)

            return HttpResponse(mensaje, content_type="text/plain")



        if "PPP" == action:
            try:
                workbook = load_workbook(
                    filename=plantilla_excel,
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.active
            except Exception as exc:
                return HttpResponseBadRequest(
                    f"No fue posible leer el archivo Excel: {exc}"
                )

            encabezados_requeridos = {
                "CONTROL VEHICULAR",
                "DESCRIPCION",
                "TOTAL (SIN IMPUESTOS)",
                "CARTAPORTE",
            }

            filas = worksheet.iter_rows(values_only=True)

            try:
                primera_fila = next(filas)
            except StopIteration:
                return HttpResponseBadRequest("El archivo está vacío.")

            encabezados = {
                str(valor).strip().upper(): indice
                for indice, valor in enumerate(primera_fila)
                if valor is not None
            }

            encabezados_faltantes = encabezados_requeridos - set(encabezados.keys())

            if encabezados_faltantes:
                return HttpResponseBadRequest(
                    "Faltan los siguientes encabezados: "
                    + ", ".join(sorted(encabezados_faltantes))
                )

            try:
                retencion = FacturapiTax.objects.get(name="RETENCIÓN")
                traslado = FacturapiTax.objects.get(name="TRASLADADO")
            except FacturapiTax.DoesNotExist as exc:
                return HttpResponseBadRequest(
                    f"No se encontró uno de los impuestos requeridos: {exc}"
                )

            productos_creados = 0
            filas_omitidas = []
            errores = []

            with transaction.atomic():
                for numero_fila, fila in enumerate(filas, start=2):
                    control_vehicular = fila[
                        encabezados["CONTROL VEHICULAR"]
                    ]
                    descripcion = fila[
                        encabezados["DESCRIPCION"]
                    ]
                    total_sin_impuestos = fila[
                        encabezados["TOTAL (SIN IMPUESTOS)"]
                    ]
                    carta_porte = fila[
                        encabezados["CARTAPORTE"]
                    ]

                    # Ignorar filas completamente vacías.
                    if not any(
                            valor not in (None, "")
                            for valor in (
                                    control_vehicular,
                                    descripcion,
                                    total_sin_impuestos,
                                    carta_porte,
                            )
                    ):
                        continue

                    if not control_vehicular or not descripcion:
                        filas_omitidas.append(numero_fila)
                        errores.append(
                            f"Fila {numero_fila}: falta CONTROL VEHICULAR "
                            "o DESCRIPCION."
                        )
                        continue

                    try:
                        precio = convertir_decimal(total_sin_impuestos)
                    except (InvalidOperation, ValueError, TypeError):
                        filas_omitidas.append(numero_fila)
                        errores.append(
                            f"Fila {numero_fila}: el TOTAL (SIN IMPUESTOS) "
                            f"no es válido: {total_sin_impuestos!r}."
                        )
                        continue

                    product = FacturapiProduct.objects.create(
                        name=str(descripcion).strip(),
                        sku=str(control_vehicular).strip(),
                        description=str(descripcion).strip(),
                        product_key="78101802",
                        unit_key="E48",
                        price=precio,
                    )

                    es_carta_porte = (
                            str(carta_porte).strip().upper() == "S"
                    )

                    if es_carta_porte:
                        product.taxes.add(retencion, traslado)
                    else:
                        product.taxes.add(traslado)

                    productos_creados += 1

            mensaje = (
                f"Se crearon exitosamente {productos_creados} productos."
            )

            if filas_omitidas:
                mensaje += (
                        f" Se omitieron {len(filas_omitidas)} filas: "
                        + "; ".join(errores)
                )
            return HttpResponse(mensaje)

        return HttpResponseBadRequest("Tipo de accion inactiva.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = ActionEngineForm()
        context.update({
            'form_action': self.form_action,
            'form': form,
            'form_type': self.form_type,
            'add_form_layout': getattr(form, 'layout', []),
            'add_form_fields': {name: form[name] for name in form.fields},
            'title': self.title,
            'category': self.category,
            'section': self.section,
        })
        return context

class ReportEngineView(AdminTemplateView):
    template_name = 'system_panel/reports_form.html'

    form_action = "ExecuteReportEngine"
    form_type = "vertical"
    title = "Motor de reportes"
    section = "Motor de reportes"
    category = "Reporteria"

    def post(self, request, *args, **kwargs):
        from django.utils.dateparse import parse_date
        report_type = request.POST.get("report_type")
        #start_date = parse_date(request.POST.get("fecha_inicial"))
        #end_date = parse_date(request.POST.get("fecha_final"))

        start_date = request.POST.get("fecha_inicial")
        end_date = request.POST.get("fecha_final")

        start_date = datetime.strptime(start_date, "%d/%m/%Y").date()
        end_date = datetime.strptime(end_date, "%d/%m/%Y").date()

        if not report_type or not start_date or not end_date:
            return HttpResponseBadRequest("Faltan parámetros: tipo, fecha de inicio o fecha de fin")

        if report_type == "folios":
            return report_xml_worksheet_folios_by_date2(request)
        elif report_type == "facturacion":
            return report_xml_invoices(request)
        elif report_type == "packing_asturiano":
            return report_asturiano(request)
        elif report_type == "asistencia":
            return report_attendance(request)
        elif report_type == "operations_master":
            return report_operations_master(request)
        elif report_type == "bitacora_cambio_turno":
            return report_operations_coordinators(request)

        return HttpResponseBadRequest("Tipo de reporte no reconocido.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report_form = ReportEngineForm()
        report_form.fields["fecha_inicial"].initial = today()
        report_form.fields["fecha_final"].initial = today() + dateutil.relativedelta.relativedelta(days=1)
        form = report_form
        context.update({
            'form_action': self.form_action,
            'form': form,
            'form_type': self.form_type,
            'add_form_layout': getattr(form, 'layout', []),
            'add_form_fields': {name: form[name] for name in form.fields},
            'title': self.title,
            'category': self.category,
            'section': self.section,
        })
        return context

class ReportEngineByFolioView(ReportEngineView):
    template_name = 'system_panel/reports_form.html'

    form_action = "ExecuteReportEngine"
    form_type = "vertical"
    title = "Motor de reportes por folio"
    section = "Motor de reportes por folio"
    category = "Reporteria"

    def post(self, request, *args, **kwargs):
        print(request.POST)
        report_type = request.POST.get("report_type")

        if report_type == "folios":
            return report_xml_worksheet_folios_by_folio(request)
        elif report_type == "facturacion":
            return report_xml_invoices(request)

        return HttpResponseBadRequest("Tipo de reporte no reconocido.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report_form = ReportEngineByFolioForm()
        form = report_form
        context.update({
            'form_action': self.form_action,
            'form': form,
            'form_type': self.form_type,
            'add_form_layout': getattr(form, 'layout', []),
            'add_form_fields': {name: form[name] for name in form.fields},
            'title': self.title,
            'category': self.category,
            'section': self.section,
        })
        return context


class ExpedienteZipProcessorView(AdminTemplateView):
    template_name = 'system_panel/expedientes_zip.html'
    title = 'Procesar expedientes (ZIP → PDFs)'
    section = 'Utilidades de documentos'
    category = 'Sistema'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = ExpedienteZipForm()
        context.update({
            'form': form,
            'form_type': 'vertical',
            'add_form_layout': getattr(form, 'layout', []),
            'add_form_fields': {name: form[name] for name in form.fields},
            'title': self.title,
            'category': self.category,
            'section': self.section,
        })
        return context

    def post(self, request, *args, **kwargs):
        from django.http import FileResponse
        from .services.expediente_processor import process_expedientes_zip, ProcessorError, ProcessorConfig
        form = ExpedienteZipForm(request.POST, request.FILES)
        if not form.is_valid():
            # Re-render with errors
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)

        upload = form.cleaned_data['zip_file']
        try:
            output_io, out_name = process_expedientes_zip(upload, config=ProcessorConfig())
        except ProcessorError as e:
            import logging
            logging.getLogger(__name__).error("Error procesando ZIP de expedientes", exc_info=True)
            form.add_error('zip_file', str(e))
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)

        response = FileResponse(output_io, as_attachment=True, filename=out_name)
        response['Content-Type'] = 'application/zip'
        return response


class PdfZipReducerView(AdminTemplateView):
    template_name = 'system_panel/expedientes_reduce_zip.html'
    title = 'Reducir PDFs (ZIP → ZIP)'
    section = 'Utilidades de documentos'
    category = 'Sistema'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = PdfReduceZipForm()
        context.update({
            'form': form,
            'form_type': 'vertical',
            'add_form_layout': getattr(form, 'layout', []),
            'add_form_fields': {name: form[name] for name in form.fields},
            'title': self.title,
            'category': self.category,
            'section': self.section,
        })
        return context

    def post(self, request, *args, **kwargs):
        from django.http import FileResponse
        from .services.pdf_reducer import process_zip_pdf_reduction
        from .services.expediente_processor import ProcessorError

        form = PdfReduceZipForm(request.POST, request.FILES)
        if not form.is_valid():
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)

        upload = form.cleaned_data['zip_file']
        quality = form.cleaned_data['quality']
        scale = form.cleaned_data['scale']
        keep_first = form.cleaned_data.get('keep_first_page', True)

        try:
            output_io, out_name = process_zip_pdf_reduction(
                upload,
                quality=quality,
                scale=scale,
                keep_first=keep_first,
            )
        except ProcessorError as e:
            import logging
            logging.getLogger(__name__).error("Error reduciendo PDFs desde ZIP", exc_info=True)
            form.add_error('zip_file', str(e))
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)

        # Force standard output name per spec
        response = FileResponse(output_io, as_attachment=True, filename='expedientes_reducidos.zip')
        response['Content-Type'] = 'application/zip'
        return response


def report_operations_master(request):
    # Build an Excel (XLSX) report of OperationMasterControl within a date range
    # Dates come as dd/mm/YYYY from the ReportEngineView form
    start_date = request.POST.get("fecha_inicial")
    end_date = request.POST.get("fecha_final")
    try:
        fecha_inicio = datetime.strptime(start_date, "%d/%m/%Y").date()
        fecha_fin = datetime.strptime(end_date, "%d/%m/%Y").date()
    except Exception:
        return HttpResponseBadRequest("Fechas inválidas. Formato esperado: dd/mm/YYYY")

    qs = (
        OperationMasterControl.objects
        .select_related(
            "operation",
            "operation__client",
            "operation__supplier",
            "operation__route",
            "operation__vehicle",
            "operation__shipment_invoice",
        )
        .prefetch_related(
            Prefetch(
                "operation__purchaseorderoperation_set",
                queryset=PurchaseOrderOperation.objects.select_related(
                    "purchase_order",
                    "purchase_order__supplier",
                ),
            ),
            "operation__invoices",
        )
        .filter(operation__operation_date__range=[fecha_inicio, fecha_fin])
        .order_by("operation__operation_date", "operation__folio", "id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Control Maestro"

    # Styles
    thin = Side(border_style=BORDER_THIN, color="CCCCCC")
    header_fill = PatternFill(start_color="F1F5FB", end_color="F1F5FB", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    headers = [
        # 1-3: Pago/VoBo/Contrarrecibo
        "Fecha a pago", "Codigo de plantilla",
        # 4-6: Datos operativos base
        "Fecha", "Código V", "Cód. Factura",
        # 7-12: Cliente/servicio/ruta/unidad
        "Fecha de factura (cliente)", "Tipo de servicio", "Cliente", "Origen", "Destino", "Unidad", "Maniobras",
        # 13-16: Ventas y cobranza
        "Precio", "Pago de clientes", "Por cobrar", "Fecha cobro (estimada)",
        # 17-21: Costos y pagos a proveedor
        "Costo", "Pagado", "Por pagar", "Proveedor", "Fecha de pago (proveedor)",
        # 22-25: Factura/OC proveedor
        "Fecha factura (proveedor)", "No. factura proveedor", "Fecha programada de pago", "Orden de compra",
        # 26-28: Factoraje
        "Monto factor", "Factoraje", "% factor",
        # 29-30: Rentabilidad
        "Utilidad", "Utilidad %",
    ]
    ws.append(headers)

    # Apply header style
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Helpers
    def safe_str(x):
        try:
            return str(x) if x is not None else ""
        except Exception:
            return ""

    # Write rows
    for c in qs:
        op = c.operation
        date_val = getattr(op, "operation_date", None)
        folio = getattr(op, "folio", None)
        client = getattr(op.client, "name", None) if op and op.client else None
        origin = destination = None
        maniobras = getattr(op, "handling_amount", None)
        if op and getattr(op, "route", None):
            origin = safe_str(getattr(op.route, "initial_location", None))
            destination = safe_str(getattr(op.route, "destination_location", None))

        if op.raw_payload:
            unit_display = safe_str(str(op.vehicle_type) if op.vehicle_type else op.raw_payload.get("unidad", ""))

        # Determine client invoice (prefer shipment_invoice, fallback to any in operation.invoices)
        invoice = None
        try:
            if getattr(op, "shipment_invoice", None):
                invoice = op.shipment_invoice
            elif hasattr(op, "invoices"):
                inv_qs = list(op.invoices.all())
                if inv_qs:
                    # Prefer ones with stamp_date, pick latest; else first
                    stamped = [i for i in inv_qs if getattr(i, "stamp_date", None)]
                    if stamped:
                        invoice = sorted(stamped, key=lambda i: i.stamp_date, reverse=True)[0]
                    else:
                        invoice = inv_qs[0]
        except Exception:
            invoice = None

        # Build invoice fields
        invoice_code = ""
        invoice_date = None
        invoice_total = None
        if invoice is not None:
            try:
                series = getattr(invoice, "series", None)
                folio_number = getattr(invoice, "folio_number", None)
                if series and folio_number is not None:
                    invoice_code = f"{series}-{folio_number}"
                elif folio_number is not None:
                    invoice_code = str(folio_number)
                elif getattr(invoice, "uuid", None):
                    invoice_code = getattr(invoice, "uuid")
                elif getattr(invoice, "facturapi_id", None):
                    invoice_code = getattr(invoice, "facturapi_id")
                # Date preference: stamp_date.date() else None
                stamp_dt = getattr(invoice, "stamp_date", None)
                if stamp_dt:
                    try:
                        invoice_date = stamp_dt.date()
                    except Exception:
                        invoice_date = None
                invoice_total = getattr(invoice, "total", None)
            except Exception:
                pass

        # Prefetched Purchase Orders for this operation
        po_links = list(getattr(op, "purchaseorderoperation_set", []).all()) if hasattr(op, "purchaseorderoperation_set") else []
        po_list = [link.purchase_order for link in po_links if getattr(link, "purchase_order", None)]

        # Compute supplier name from Operation or first PO
        supplier_name = None
        if op and getattr(op, "supplier", None) and getattr(op.supplier, "name", None):
            supplier_name = op.supplier.name
        elif po_list:
            po_sup = getattr(po_list[0], "supplier", None)
            supplier_name = getattr(po_sup, "business_name", None) or getattr(po_sup, "name", None)

        # Aggregate PO folios
        po_folios = ", ".join(sorted({safe_str(getattr(po, "folio", "")) for po in po_list if getattr(po, "folio", None)})) or (c.purchase_order or "")

        # Numeric conversions for Excel (keep as numbers)
        def n(val):
            if val is None:
                return 0
            try:
                return float(val)
            except Exception:
                try:
                    return float(str(val).replace(",", "."))
                except Exception:
                    return 0

        price = n(c.sale_amount)

        # Cost from Purchase Orders: base on this operation's total plus its accessories in each PO
        # If there are no POs, fallback to control.cost_amount
        if po_list:
            op_base_total = n(getattr(op, "total", 0))
            accessories_sum = 0.0
            for po in po_list:
                # accessories are prefetched; filter by this operation
                try:
                    accs = list(po.accessories.all())
                except Exception:
                    accs = []
                for acc in accs:
                    try:
                        if getattr(acc, "operation_id", None) == getattr(op, "id", None):
                            accessories_sum += n(getattr(acc, "subtotal", 0))
                    except Exception:
                        continue
            cost = op_base_total + accessories_sum
        else:
            cost = n(c.cost_amount)

        # Supplier paid and dates from PO
        paid_supplier = 0.0
        supplier_paid_date = None
        if po_list:
            # Consider paid if any PO is in status PAGADA or has paid_date
            any_paid = False
            latest_paid_dt = None
            for po in po_list:
                status = getattr(po, "status", None)
                paid_dt = getattr(po, "paid_date", None)
                if status == PurchaseOrderStatus.PAGADA or paid_dt:
                    any_paid = True
                    if paid_dt and (latest_paid_dt is None or paid_dt > latest_paid_dt):
                        latest_paid_dt = paid_dt
            if any_paid:
                paid_supplier = cost  # assume fully paid when PO shows as paid
            supplier_paid_date = latest_paid_dt.date() if latest_paid_dt else None

        pending_supplier = max(0.0, cost - paid_supplier)

        # Client collection placeholders (Fase 2)
        paid_clients = 0.0
        pending_clients = max(0.0, price - paid_clients)

        factoring_cost = n(c.factoring_cost)
        profit = n(c.profit)
        # Excel expects fractions (0-1) for percentage cells
        profit_pct_fraction = (n(c.profit_percentage) / 100.0) if price else 0.0
        factoring_pct_fraction = n(c.factoring_percentage) / 100.0

        # Scheduled supplier payment date: try PO.approved_date date; fallback to control.scheduled_supplier_payment_date
        scheduled_supplier_payment = None
        if po_list:
            # choose the earliest approved_date among POs for scheduling reference
            dates = [po.approved_date.date() for po in po_list if getattr(po, "approved_date", None)]
            if dates:
                scheduled_supplier_payment = min(dates)
        if not scheduled_supplier_payment:
            scheduled_supplier_payment = c.scheduled_supplier_payment_date

        # Supplier invoice date: no explicit field in PO model; fallback to control.supplier_invoice_date
        supplier_invoice_date = c.supplier_invoice_date

        # Determine price to use: prefer invoice total when available
        price_final = n(invoice_total) if (invoice_total is not None) else price

        row = [
            # 1-3: Pago/VoBo/Contrarrecibo
            c.counter_receipt_date,
            c.counter_receipt or "",
            # 4-6
            date_val,
            folio or "",
            invoice_code or "",
            # 7-12
            invoice_date,
            "TRANSLADO",
            client or "",
            origin or "",
            destination or "",
            unit_display or "",
            maniobras,
            # 13-16
            price_final,
            paid_clients,
            pending_clients if invoice_total is None else max(0.0, price_final - paid_clients),
            c.expected_collection_date,
            # 17-21
            cost,
            paid_supplier,
            pending_supplier,
            supplier_name or "",
            supplier_paid_date,
            # 22-25
            supplier_invoice_date,
            c.supplier_invoice_number or "",
            scheduled_supplier_payment,
            po_folios,
            # 26-28
            n(c.factoring_amount),
            "Sí" if c.has_factoring else "No",
            factoring_pct_fraction,
            # 29-30
            profit if invoice_total is None else (price_final - cost - factoring_cost),
            profit_pct_fraction if invoice_total is None else ((price_final - cost - factoring_cost) / price_final if price_final else 0.0),
        ]
        ws.append(row)

    # Number formats and alignment for columns
    money_fmt = '"$"#,##0.00'
    percent_fmt = '0.00%'
    # Indices basados en el nuevo orden de encabezados (1-based):
    # 13 Precio, 14 Pago clientes, 15 Por cobrar, 17 Costo, 18 Pagado, 19 Por pagar, 26 Monto factor, 29 Utilidad
    money_cols = [13, 14, 15, 17, 18, 19, 26, 29]
    # 28 % factor, 30 Utilidad %
    percent_cols = [28, 30]

    for r in range(2, ws.max_row + 1):
        for cidx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=cidx)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cidx in money_cols:
                cell.number_format = money_fmt
                cell.alignment = right
            elif cidx in percent_cols:
                cell.number_format = percent_fmt
                cell.alignment = right

    # Auto width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    length = 0
                elif isinstance(val, (int, float)):
                    length = len(str(val))
                else:
                    length = len(str(val))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_length + 2), 45)

    # Freeze header
    ws.freeze_panes = "A2"

    # Build response
    filename = f"control_maestro_{fecha_inicio.isoformat()}_{fecha_fin.isoformat()}.xlsx"
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f"attachment; filename={filename}"
    return resp

def report_asturiano(request):
    start_date = request.POST.get("fecha_inicial")
    end_date = request.POST.get("fecha_final")
    fecha_inicio = datetime.strptime(start_date, "%d/%m/%Y").date()
    fecha_fin = datetime.strptime(end_date, "%d/%m/%Y").date()
    operations = Operation.objects.filter(
        operation_date__range=[fecha_inicio, fecha_fin],

    ).order_by("operation_date")
    wb = Workbook()
    ws = None
    date_title = ""
    base_col = 0
    clients = [
        Client.objects.get(name__icontains="Asturiano"),
        Client.objects.get(name__icontains="Asturito"),
    ]
    print(clients)
    print(fecha_inicio)
    print(fecha_fin)
    print(Operation.objects.filter(operation_date=fecha_inicio))
    operations = Operation.objects.filter(
        operation_date__range=[fecha_inicio, fecha_fin],
    ).order_by("operation_date")

    print("fecha_inicio:", fecha_inicio)
    print("fecha_fin:", fecha_fin)
    print("en fecha_inicio:", Operation.objects.filter(operation_date=fecha_inicio).count())
    print("en rango:", operations.count())
    print(list(operations.values("id", "operation_date")))
    for operation in Operation.objects.filter(client__in=clients).filter(operation_date__range=[fecha_inicio, fecha_fin]).order_by("operation_date").all():
        print(operation)
        if date_title != str(operation.operation_date):
            date_title = str(operation.operation_date)
            ws = wb.create_sheet(operation.folio)
            ws.title = date_title
            base_col = 0
        else:
            base_col += 8

        ws['A' + str(base_col + 1)] = operation.folio
        ws['A' + str(base_col + 2)] = 'CLIENTE'
        ws['A' + str(base_col + 3)] = 'OPERADOR'
        ws['A' + str(base_col + 4)] = 'PLACAS'
        ws['A' + str(base_col + 5)] = 'RUTA'
        ws['C' + str(base_col + 4)] = 'ECONOMICO'
        ws['C' + str(base_col + 5)] = 'FOLIO INTERNO DE CARGA'
        ws['E' + str(base_col + 4)] = 'PROVEEDOR'

        ws['A' + str(base_col + 6)] = 'CANTIDAD'
        ws['B' + str(base_col + 6)] = 'DESCRIPCION'
        ws['C' + str(base_col + 6)] = 'PESO'
        ws['D' + str(base_col + 6)] = 'CLAVE'
        ws['E' + str(base_col + 6)] = 'TIENDA'
        ws['F' + str(base_col + 6)] = 'CLASIF'

        ws['B' + str(base_col + 2)] = str(operation.client)
        ws['B' + str(base_col + 3)] = str(operation.driver)
        if operation.vehicle:
            ws['B' + str(base_col + 4)] = str(operation.vehicle.license_plate)
            ws['B' + str(base_col + 5)] = str("")
            ws['D' + str(base_col + 4)] = str(operation.vehicle)
        ws['D' + str(base_col + 5)] = ""
        ws['E' + str(base_col + 5)] = str(operation.driver)

        ws.merge_cells(start_row=base_col + 1, start_column=1, end_row=base_col + 1, end_column=6)
        ws.merge_cells(start_row=base_col + 2, start_column=2, end_row=base_col + 2, end_column=6)
        ws.merge_cells(start_row=base_col + 3, start_column=2, end_row=base_col + 3, end_column=6)
        ws.merge_cells(start_row=base_col + 4, start_column=5, end_row=base_col + 4, end_column=6)
        ws.merge_cells(start_row=base_col + 5, start_column=5, end_row=base_col + 5, end_column=6)

        titleFill = PatternFill(start_color='EBA67D',
                                end_color='EBA67D',
                                fill_type='solid',
                                )
        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )

        ws['A' + str(base_col + 1)].fill = titleFill
        ws['A' + str(base_col + 2)].fill = titleFill
        ws['A' + str(base_col + 3)].fill = titleFill
        ws['A' + str(base_col + 4)].fill = titleFill
        ws['A' + str(base_col + 5)].fill = titleFill
        ws['C' + str(base_col + 4)].fill = titleFill
        ws['C' + str(base_col + 5)].fill = titleFill
        ws['E' + str(base_col + 4)].fill = titleFill

        ws['A' + str(base_col + 6)].fill = titleFill
        ws['B' + str(base_col + 6)].fill = titleFill
        ws['C' + str(base_col + 6)].fill = titleFill
        ws['D' + str(base_col + 6)].fill = titleFill
        ws['E' + str(base_col + 6)].fill = titleFill
        ws['F' + str(base_col + 6)].fill = titleFill

        controlador = base_col + 7

        # for transportedProduct in operation.transported_products.all():
        #     # Cantidad
        #     ws.cell(row=controlador, column=1).value = str(transportedProduct.amount)
        #     # Descripcion
        #     ws.cell(row=controlador, column=2).value = str(transportedProduct.description)
        #     # Peso
        #     ws.cell(row=controlador, column=3).value = str(transportedProduct.weight)
        #     # Clave
        #     ws.cell(row=controlador, column=4).value = str(transportedProduct.transported_product_key)
        #     # Destino
        #     if operation.route:
        #         ws.cell(row=controlador, column=5).value = str(operation.route.destination_location)
        #     # Clasif


        for packing in DistributionPacking.objects.filter(operation=operation).all():
            ws.cell(row=controlador, column=1).value = str(packing.amount)
            ws.cell(row=controlador, column=2).value = str(packing.distribution)
            ws.cell(row=controlador, column=3).value = str(packing.weight)
            if packing.delivery_shop:
                ws.cell(row=controlador, column=4).value = str(packing.delivery_shop.name.split(' ')[0])
                ws.cell(row=controlador, column=5).value = str(packing.delivery_shop.name)
            controlador += 1
            base_col += 1

        dims = {}
        i = 1
        for row in ws.rows:
            j = 1
            for cell in row:
                ws.cell(row=i, column=j).border = thin_border
                ws.cell(row=i, column=j).alignment = Alignment(horizontal='center')
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 10))
                j += 1
            i += 1
        for col, value in dims.items():
            ws.column_dimensions[col].width = 20

    # Establecer el nombre de mi archivo
    nombre_archivo = "Packing Asturiano " + str(fecha_inicio) + " - " + str(fecha_fin) +".xlsx"
    # Definir el tipo de respuesta que se va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

def convertir_decimal(valor):
    """
    Convierte valores como:
    1500
    1500.50
    $1,500.50
    '1,500.50'
    a Decimal.
    """
    if valor is None or valor == "":
        raise ValueError("El valor está vacío.")

    if isinstance(valor, Decimal):
        return valor

    if isinstance(valor, (int, float)):
        return Decimal(str(valor))

    valor_limpio = (
        str(valor)
        .strip()
        .replace("$", "")
        .replace(",", "")
        .replace(" ", "")
    )

    return Decimal(valor_limpio)


def report_operations_coordinators(request):
    start_date = request.POST.get("fecha_inicial")
    end_date = request.POST.get("fecha_final")

    try:
        fecha_inicio = datetime.strptime(start_date, "%d/%m/%Y").date()
        fecha_fin = datetime.strptime(end_date, "%d/%m/%Y").date()
    except Exception:
        return HttpResponseBadRequest(
            "Fechas inválidas. Formato esperado: dd/mm/YYYY"
        )

    operations = (
        Operation.objects
        .filter(operation_date__range=[fecha_inicio, fecha_fin])
        .order_by("operation_date", "folio")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Operaciones"

    def style_section(cell_range, fill):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill
                cell.font = main_header_font
                cell.alignment = center
                cell.border = thin_border

    # =========================================================
    # ESTILOS
    # =========================================================

    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color="000000"),
        right=Side(border_style=BORDER_THIN, color="000000"),
        top=Side(border_style=BORDER_THIN, color="000000"),
        bottom=Side(border_style=BORDER_THIN, color="000000"),
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    top_center = Alignment(
        horizontal="center",
        vertical="top",
        wrap_text=True,
    )

    # Texto de encabezados principales
    main_header_font = Font(
        color="FFFFFF",
        bold=True,
        size=11,
    )

    # Identificación del servicio - Azul marino
    identification_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    # Planeación - Naranja
    planning_fill = PatternFill(
        start_color="C65911",
        end_color="C65911",
        fill_type="solid",
    )

    # 1er turno - Verde
    first_shift_fill = PatternFill(
        start_color="548235",
        end_color="548235",
        fill_type="solid",
    )

    # 2do turno - Azul
    second_shift_fill = PatternFill(
        start_color="2F75B5",
        end_color="2F75B5",
        fill_type="solid",
    )

    # 3er turno - Morado
    third_shift_fill = PatternFill(
        start_color="7030A0",
        end_color="7030A0",
        fill_type="solid",
    )

    # Jefatura - Rojo oscuro
    management_fill = PatternFill(
        start_color="C00000",
        end_color="C00000",
        fill_type="solid",
    )

    # Encabezados de columnas
    header_fill = PatternFill(
        start_color="D9E1F2",
        end_color="D9E1F2",
        fill_type="solid",
    )

    # =========================================================
    # ENCABEZADOS PRINCIPALES
    # =========================================================

    # Identificación del servicio
    ws.merge_cells("A1:L1")
    ws.merge_cells("M1:Q1")
    ws.merge_cells("R1:AK1")
    ws.merge_cells("AL1:BE1")
    ws.merge_cells("BF1:BY1")
    ws["A1"] = "Identificación del servicio"

    # Planeación
    ws["A1"] = "Identificación del servicio"
    ws["M1"] = "Planeación"
    # 1er turno
    #ws.merge_cells("D1:AK1")
    ws["R1"] = "1er. Turno ( 8 am a 3 pm) Coordinadores"
    ws["AL1"] = "2do. Turno (3 pm a 10 pm) Monitoreo Operativo"
    ws["BF1"] = "3er.Turno (10 pm a 6 am) Monitoreo Nocturno"
    ws["BZ1"] = "Jefatura"

    # Colores de encabezados principales
    for row in ws["A1:B1"]:
        for cell in row:
            cell.fill = identification_fill
            cell.alignment = center
            cell.border = thin_border

    ws["C1"].fill = planning_fill
    ws["C1"].alignment = center
    ws["C1"].border = thin_border

    style_section("A1:L1", identification_fill)
    style_section("M1:Q1", planning_fill)
    style_section("R1:AK1", first_shift_fill)
    style_section("AL1:BE1", second_shift_fill)
    style_section("BF1:BY1", third_shift_fill)

    # Jefatura por ahora es solamente BZ
    style_section("BZ1:BZ1", management_fill)

    for row in ws["D1:BZ1"]:
        for cell in row:
            cell.alignment = center
            cell.border = thin_border

    # =========================================================
    # ENCABEZADOS DE COLUMNAS
    # =========================================================

    headers = [
        "Fecha",
        "C. Vehicular",
        "Cliente",
        "Coordinador",
        "Origen",
        "Destino",
        "Repartos",
        "Proveedor",
        "Operador",
        "Placas",
        "Economico",
        "Division",
        "C. carga",
        "H. de salida",
        "C. descarga",
        "Ruta",
        "Indicaciones de seguridad",

        "H. llegada",
        "Semaforo",
        "Motivo",
        "H. de salida",
        "Evidencias Carga",
        "Carta porte",
        "Estatus (Corte 11 am)",
        "Comentarios",
        "Estatus (Corte 2 pm)",
        "Comentarios",
        "Incidencias",
        "Comentarios",
        "Dadivas",
        "Comentarios",
        "Evidencias (pod's clientes)",
        "Evidencias (carga en drive)",
        "Estatus cambio de turno",
        "Pendientes",
        "Calificacion (0 al 10 pts)",
        "Motivo",

        "H. llegada",
        "Semaforo",
        "Motivo",
        "H. de salida",
        "Evidencias Carga",
        "Carta porte",
        "Estatus (Corte 11 am)",
        "Comentarios",
        "Estatus (Corte 2 pm)",
        "Comentarios",
        "Incidencias",
        "Comentarios",
        "Dadivas",
        "Comentarios",
        "Evidencias (pod's clientes)",
        "Evidencias (carga en drive)",
        "Estatus cambio de turno",
        "Pendientes",
        "Calificacion (0 al 10 pts)",
        "Motivo",

        "H. llegada",
        "Semaforo",
        "Motivo",
        "H. de salida",
        "Evidencias Carga",
        "Carta porte",
        "Estatus (Corte 11 am)",
        "Comentarios",
        "Estatus (Corte 2 pm)",
        "Comentarios",
        "Incidencias",
        "Comentarios",
        "Dadivas",
        "Comentarios",
        "Evidencias (pod's clientes)",
        "Evidencias (carga en drive)",
        "Estatus cambio de turno",
        "Pendientes",
        "Calificacion (0 al 10 pts)",
        "Motivo",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(
            row=2,
            column=col_idx,
            value=header,
        )

        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # =========================================================
    # DATOS
    # =========================================================

    row_number = 3

    for operation in operations:
        row = [
            # A - Fecha
            operation.operation_date,

            # B - C. Vehicular
            operation.folio,

            # C - Cliente
            str(operation.client),

            # D - Coordinador
            "",

            # E - Origen
            str(operation.route.initial_location) if operation.route else "",

            # F - Destino
            str(operation.route.destination_location) if operation.route else "",

            # G - Repartos
            "",

            # H - Proveedor
            str(operation.supplier),

            # I - Operador
            str(operation.driver),

            # J - Placas
            operation.vehicle.license_plate if operation.vehicle else "",

            # K - Economico
            "",

            # L - Division
            "",

            # M - C. carga
            excel_datetime(operation.cargo_appointment),

            # N - H. de salida
            excel_datetime(operation.scheduled_departure_time),

            # O - C. descarga
            excel_datetime(operation.download_appointment),

            # P - Ruta
            "",

            # Q - Indicaciones de seguridad
            "",

            # R - H. llegada
            "",

            # S - Semaforo
            "",

            # T - Motivo
            "",

            # U - H. de salida
            "",

            # V - Evidencias Carga
            "",

            # W - Carta porte
            "",

            # X - Estatus corte 11 am
            "",

            # Y - Comentarios
            "",

            # Z - Estatus corte 2 pm
            "",

            # AA - Comentarios
            "",

            # AB - Incidencias
            "",

            # AC - Comentarios
            "",

            # AD - Dadivas
            "",

            # AE - Comentarios
            "",

            # AF - Evidencias pod's clientes
            "",

            # AG - Evidencias carga drive
            "",

            # AH - Estatus cambio de turno
            "",

            # AI - Pendientes
            "",

            # AJ - Calificación
            "",

            # AK - Motivo
            "",

            # R - H. llegada
            "",

            # S - Semaforo
            "",

            # T - Motivo
            "",

            # U - H. de salida
            "",

            # V - Evidencias Carga
            "",

            # W - Carta porte
            "",

            # X - Estatus corte 11 am
            "",

            # Y - Comentarios
            "",

            # Z - Estatus corte 2 pm
            "",

            # AA - Comentarios
            "",

            # AB - Incidencias
            "",

            # AC - Comentarios
            "",

            # AD - Dadivas
            "",

            # AE - Comentarios
            "",

            # AF - Evidencias pod's clientes
            "",

            # AG - Evidencias carga drive
            "",

            # AH - Estatus cambio de turno
            "",

            # AI - Pendientes
            "",

            # AJ - Calificación
            "",

            # AK - Motivo
            "",

            # R - H. llegada
            "",

            # S - Semaforo
            "",

            # T - Motivo
            "",

            # U - H. de salida
            "",

            # V - Evidencias Carga
            "",

            # W - Carta porte
            "",

            # X - Estatus corte 11 am
            "",

            # Y - Comentarios
            "",

            # Z - Estatus corte 2 pm
            "",

            # AA - Comentarios
            "",

            # AB - Incidencias
            "",

            # AC - Comentarios
            "",

            # AD - Dadivas
            "",

            # AE - Comentarios
            "",

            # AF - Evidencias pod's clientes
            "",

            # AG - Evidencias carga drive
            "",

            # AH - Estatus cambio de turno
            "",

            # AI - Pendientes
            "",

            # AJ - Calificación
            "",

            # AK - Motivo
            "",
        ]

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(
                row=row_number,
                column=col_idx,
                value=value,
            )

            cell.border = thin_border
            cell.alignment = top_center

        row_number += 1

    # =========================================================
    # FORMATOS
    # =========================================================

    # Fecha
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = "dd/mm/yyyy"

    # Horas
    time_columns = [
        14,  # N - H. salida
        18,  # R - H. llegada
        21,  # U - H. salida
    ]

    for row in range(3, ws.max_row + 1):
        for col_idx in time_columns:
            ws.cell(
                row=row,
                column=col_idx,
            ).number_format = "hh:mm"

    # =========================================================
    # TAMAÑOS
    # =========================================================

    widths = {
        "A": 13,
        "B": 16,
        "C": 25,
        "D": 25,
        "E": 25,
        "F": 25,
        "G": 12,
        "H": 25,
        "I": 25,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 25,
        "Q": 35,
        "R": 15,
        "S": 15,
        "T": 25,
        "U": 15,
        "V": 25,
        "W": 20,
        "X": 22,
        "Y": 30,
        "Z": 22,
        "AA": 30,
        "AB": 25,
        "AC": 30,
        "AD": 15,
        "AE": 30,
        "AF": 30,
        "AG": 30,
        "AH": 25,
        "AI": 30,
        "AJ": 22,
        "AK": 30,

        "AL": 15,
        "AM": 15,
        "AN": 25,
        "AO": 15,
        "AP": 25,
        "AQ": 20,
        "AR": 22,
        "AS": 30,
        "AT": 22,
        "AU": 30,
        "AV": 25,
        "AW": 30,
        "AX": 15,
        "AY": 30,
        "AZ": 30,
        "BA": 30,
        "BB": 25,
        "BC": 30,
        "BD": 22,
        "BE": 30,

        "BF": 15,
        "BG": 15,
        "BH": 25,
        "BI": 15,
        "BJ": 25,
        "BK": 20,
        "BL": 22,
        "BM": 30,
        "BN": 22,
        "BO": 30,
        "BP": 25,
        "BQ": 30,
        "BR": 15,
        "BS": 30,
        "BT": 30,
        "BU": 30,
        "BV": 25,
        "BW": 30,
        "BX": 22,
        "BY": 30,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 55

    # =========================================================
    # CONFIGURACIÓN DE LA HOJA
    # =========================================================

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:AK{max(ws.max_row, 2)}"

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # =========================================================
    # RESPUESTA
    # =========================================================

    nombre_archivo = (
        f"Operaciones Coordinadores "
        f"{fecha_inicio} - {fecha_fin}.xlsx"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    contenido = (
        'attachment; filename="{0}"'
        .format(nombre_archivo)
    )

    response["Content-Disposition"] = contenido

    wb.save(response)

    return response


def excel_datetime(value):
    """
    Convierte un datetime de Django a un datetime compatible con Excel.

    - Conserva la hora local.
    - Elimina tzinfo porque Excel/openpyxl no soporta timezones.
    """
    if value is None:
        return None

    if isinstance(value, datetime) and timezone.is_aware(value):
        value = timezone.localtime(value)
        value = value.replace(tzinfo=None)

    return value.time() if isinstance(value, datetime) else value

